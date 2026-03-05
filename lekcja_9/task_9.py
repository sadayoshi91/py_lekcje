# Używając openpyxl , stwórz plik finanse.xlsx . W
# pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich
# wartości. W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając
# formuły Excela (np. =SUM(B1:B2) ).

from openpyxl import Workbook

def create_finance_file(filename):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Finanse"
    expenses = [("Czynsz", 1500), ("Jedzenie", 800)]
    for row in expenses:
        ws.append(row)
    ws["B3"] = "=SUM(B1:B2)"
    wb.save(filename)
filename = "lekcja_9/finanse.xlsx"
create_finance_file(filename)
